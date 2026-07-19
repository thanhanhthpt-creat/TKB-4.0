from __future__ import annotations

from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook


class ExcelService:
    REQUIRED_EXTENSIONS = {".xlsx", ".xlsm"}

    @staticmethod
    def inspect_workbook(path: Path) -> list[str]:
        if not path.exists():
            raise FileNotFoundError(f"Không tìm thấy tệp Excel: {path}")
        if path.suffix.lower() not in ExcelService.REQUIRED_EXTENSIONS:
            raise ValueError("Khung thời khóa biểu phải là tệp .xlsx hoặc .xlsm")

        wb = load_workbook(path, read_only=True, data_only=False)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()

    @staticmethod
    def export_copy(source: Path, destination: Path) -> Path:
        destination.parent.mkdir(parents=True, exist_ok=True)
        copy2(source, destination)
        return destination
